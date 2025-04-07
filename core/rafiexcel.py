'''
RafiExcel Class for Excel Manipulation
'''

from typing import List, Optional
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from starlette.datastructures import UploadFile
from tempfile import NamedTemporaryFile
import io
from PIL import Image as PILImage
import os
from core.file import download_file_to_bytes
from datetime import datetime, timedelta

# Color Definition
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
blue_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")

# Border Definition
thick_border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
thin_border_up = Border(
    top=Side(style='thin'),
)

class RafiExcel:
    def __init__(self):
        pass

    def merge_and_center_text(
            self, 
            column:str, 
            range_col:str, 
            ws:any, 
            title:str, 
            color:Optional[PatternFill] = None,
            ):
        if color != None:
            ws[column].fill = color
        ws[column] = title
        ws[column].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(range_col)
        return ws
    def merge_and_left_text(
            self, 
            column:str, 
            range_col:str, 
            ws:any, 
            title:str, 
            color:Optional[PatternFill] = None
            ):
        if color != None:
            ws[column].fill = color
        ws[column] = title
        ws[column].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(range_col)
        return ws
    def merge_and_right_text(
            self, 
            column:str, 
            range_col:str, 
            ws:any, 
            title:str, 
            color:Optional[PatternFill] = None
            ):
        if color != None:
            ws[column].fill = color
        ws[column] = title
        ws[column].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(range_col)
        return ws
    def get_col_name(
            self, 
            nomor_kolom, 
            nomor_baris
            ):
        nama_kolom = ''
        while nomor_kolom > 0:
            nomor_kolom, sisa = divmod(nomor_kolom - 1, 26)
            nama_kolom = chr(65 + sisa) + nama_kolom
        referensi_sel = f"{nama_kolom}{nomor_baris}"
        return referensi_sel
    def get_col_index(
            self, 
            nomor_kolom
            ):
        nama_kolom = ''
        while nomor_kolom > 0:
            nomor_kolom, sisa = divmod(nomor_kolom - 1, 26)
            nama_kolom = chr(65 + sisa) + nama_kolom
        return nama_kolom

    def insert_gambar(
            self, 
            ws:any, 
            column:str, 
            column_id:str, 
            row:int
            ):
        current_directory = os.getcwd()
        img_path = f'{current_directory}/core/example.jpg'
        # img = Image(img_path)
        img = PILImage.open(img_path)
        ws.column_dimensions[column_id].width = 20
        ws.row_dimensions[row].height = 100

        img_width, img_height = img.size
        cell_width = ws.column_dimensions[column_id].width * 7.9
        cell_height = ws.row_dimensions[row].height * 1.33
        width_ratio = cell_width / img_width
        height_ratio = cell_height / img_height
        scale_ratio = min(width_ratio, height_ratio)
        new_width = img_width * scale_ratio
        new_height = img_height * scale_ratio
        img_resized = Image(img_path)
        img_resized.width = new_width
        img_resized.height = new_height
        # Add image
        ws.add_image(img_resized, column)
        return ws
    def insert_gambar_custom(
            self, 
            ws:any, 
            column:str, 
            column_id:str, 
            row:int, 
            path_file:str
            ):
        image_data = download_file_to_bytes(path=path_file)
        # img = PILImage.open(io.BytesIO(image_data))
        img = PILImage.open(image_data)
        ws.column_dimensions[column_id].width = 30
        ws.row_dimensions[row].height = 160
        newsize = (210, 210)
        img_resized = img.resize(newsize)
        # Simpan gambar ke dalam buffer
        img_buffer = io.BytesIO()
        img_resized.save(img_buffer, format="PNG")
        img_buffer.seek(0)
        # Add image
        excel_image = Image(img_buffer)
        ws.add_image(excel_image, column)
        return ws
    
    def tilt_text(
        self, 
        ws:any,
        text:str,
        column:str, 
        textRotation:int=55,
        wrapText:bool=True,
    ):
        ws[column] = text
        ws[column].fill = blue_fill
        ws[column].alignment = Alignment(textRotation=textRotation, wrapText=wrapText, vertical='center')
        return ws
    
    def text_center_and_color(
        self, 
        ws:any,
        text:str,
        column:str, 
        color:PatternFill,
    ):
        ws[column].fill = color
        ws[column] = text
        ws[column].alignment = Alignment(horizontal='center', wrapText=True, vertical='center')
        return ws
    
    def text_right_and_color(
        self, 
        ws:any,
        text:str,
        column:str, 
        color:PatternFill,
    ):
        ws[column].fill = color
        ws[column] = text
        ws[column].alignment = Alignment(horizontal='right', wrapText=True, vertical='center')
        return ws
    
    def text_left_and_color(
        self, 
        ws:any,
        text:str,
        column:str, 
        color:PatternFill,
    ):
        ws[column].fill = color
        ws[column] = text
        ws[column].alignment = Alignment(horizontal='left', wrapText=True, vertical='center')
        return ws
    
    def text_center(
        self, 
        ws:any,
        text:str,
        column:str,
        wrap_text: Optional[bool] = True
    ):
        ws[column] = text
        ws[column].alignment = Alignment(horizontal='center', wrapText=True, vertical='center')
        return ws
    
    def text_center_bold(
        self, 
        ws: any,
        text: str,
        column: str,
        wrap_text: Optional[bool] = True
    ):
        ws[column] = text
        ws[column].alignment = Alignment(horizontal='center', wrapText=wrap_text, vertical='center')
        ws[column].font = Font(bold=True)
        return ws
    
    def text_left(
        self, 
        ws:any,
        text:str,
        column:str, 
    ):
        ws[column] = text
        ws[column].alignment = Alignment(horizontal='left', wrapText=True, vertical='center')
        return ws
    
    def text_right(
        self, 
        ws:any,
        text:str,
        column:str, 
    ):
        ws[column] = text
        ws[column].alignment = Alignment(horizontal='right', wrapText=True, vertical='center')
        return ws
    
    def read_excel_column(self, file_content, column_name):
        with BytesIO(file_content.read()) as bytes_io:
            workbook = load_workbook(bytes_io)
            sheet = workbook.active

            column_index = None
            for cell in sheet[1]:
                if cell.value == column_name:
                    column_index = cell.column
                    break
            if column_index is None:
                raise ValueError(f"Column '{column_name}' not found in Excel file.")

            column_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                column_data.append(row[column_index - 1])

        return column_data
    def excel_to_list_of_dicts(self, file_content):
        try:
            #TODO: change to query 
            ls_execl_field=['no', 'id', 'nama', 'alamat', 'expire']
            with BytesIO(file_content.read()) as bytes_io:
                workbook = load_workbook(bytes_io)
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]
                print('headers',headers)
                if len(ls_execl_field) == len(headers):
                    validation_field = all(field in ls_execl_field for field in headers)
                    if not validation_field:
                        raise ValueError(f"Error: Missing or incorrect columns. Please check and re-upload your file.")
                else:
                    raise ValueError(f"Error: Missing or incorrect columns. Please check and re-upload your file.")
                data_list = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_data = {headers[i]: row[i] for i in range(len(headers))}
                    data_list.append(row_data)

                return data_list
        except ValueError as e:
            print('error :',e)
            return None

    def excel_date_to_datetime(excel_date: int) -> datetime:
        excel_base_date = datetime(1899, 12, 30)
        return (excel_base_date + timedelta(days=excel_date)).date()
    


# help_excel = RafiExcel()

# async def generate_excel(data:List[any], ls_pekerjaan:List[any], photo_data:List[any]):
#     wb = Workbook()
#     # Add data to wb
#     ws1 = wb.create_sheet(title='CEKLIST')
#     ws1 = help_excel.merge_and_center_text(ws=ws1,range_col='A1:X1', column='A1',color=blue_fill, title="FORM PM PERANGKAT IT")
#     # Setup with and high
#     ws1.row_dimensions[8].height = 76.5
#     ws1.column_dimensions['A'].width = 3.57
#     ws1.column_dimensions['B'].width = 17
#     ws1.column_dimensions['C'].width = 17
#     ws1.column_dimensions['D'].width = 17
#     ws1.column_dimensions['E'].width = 17
#     ws1.column_dimensions['F'].width = 11
#     ws1.column_dimensions['G'].width = 6
#     ws1.column_dimensions['H'].width = 6
#     ws1.column_dimensions['I'].width = 6
#     ws1.column_dimensions['J'].width = 6
#     ws1.column_dimensions['K'].width = 6
#     ws1.column_dimensions['L'].width = 6
#     ws1.column_dimensions['M'].width = 6
#     ws1.column_dimensions['N'].width = 6
#     ws1.column_dimensions['O'].width = 6
#     ws1.column_dimensions['P'].width = 6
#     ws1.column_dimensions['Q'].width = 20
#     ws1.column_dimensions['W'].width = 20
#     ws1.column_dimensions['X'].width = 20

#     ws1 = help_excel.text_left(ws=ws1, text='KANWIL', column='B3')
#     wilayah_data = data[0].pm.cabang.wilayah.name_wilayah if data[0].pm.cabang else ''
#     ws1 = help_excel.help_excel.text_left(ws=ws1, text=f': {wilayah_data}', column='C3')
#     ws1 = help_excel.text_left(ws=ws1, text='KANCA', column='B4')
#     cabang_data = data[0].pm.cabang.name_cabang if data[0].pm.cabang else ''
#     ws1 = help_excel.text_left(ws=ws1, text=f': {cabang_data}', column='C4')
#     ws1 = help_excel.text_left(ws=ws1, text='Nama Uker', column='B5')
#     unitkerja_data = data[0].pm.unit_kerja[0].name_unit_kerja if data[0].pm.unit_kerja else ''
#     ws1 = help_excel.text_left(ws=ws1, text=f': {unitkerja_data}', column='C5')
#     ws1 = help_excel.text_left(ws=ws1, text='Kode Uker', column='B6')
#     kode_unit_data = data[0].pm.unit_kerja[0].code_unit_kerja if data[0].pm.unit_kerja else ''
#     ws1 = help_excel.text_left(ws=ws1, text=f': {kode_unit_data}', column='C6')

#     ws1 = help_excel.text_center_and_color(ws=ws1, text='NO', column='A8', color=blue_fill)
#     ws1 = help_excel.text_center_and_color(ws=ws1, text='ASET ID', column='B8', color=blue_fill)
#     ws1 = help_excel.text_center_and_color(ws=ws1, text='SN', column='C8', color=blue_fill)
#     ws1 = help_excel.text_center_and_color(ws=ws1, text='NAMA PERANGKAT', column='D8', color=blue_fill)
#     ws1 = help_excel.text_center_and_color(ws=ws1, text='MEREK/TYPE', column='E8', color=blue_fill)
#     ws1 = help_excel.text_center_and_color(ws=ws1, text='UMUR PERANGKAT', column='F8', color=blue_fill)
#     ws1 = help_excel.text_center_and_color(ws=ws1, text='NAMA USER (Pemegang Perangkat)', column='Q8', color=blue_fill)
#     ws1 = help_excel.text_center_and_color(ws=ws1, text='JABATAN', column='R8', color=blue_fill)
#     ws1 = help_excel.text_center_and_color(ws=ws1, text='PERSONAL NUMBER (PN)', column='S8', color=blue_fill)
#     ws1 = help_excel.text_center_and_color(ws=ws1, text='KETERANGAN', column='T8', color=blue_fill)
#     ws1 = help_excel.merge_and_center_text(ws=ws1, column='U8', range_col='U8:V8', title='PENGERJAAN PM', color=blue_fill)
#     ws1 = help_excel.merge_and_center_text(ws=ws1, column='W8', range_col='W8:X8', title='EVIDENCE', color=blue_fill)

#     ws1= help_excel.text_center(ws=ws1, column='U9', text='Manual')
#     ws1= help_excel.text_center(ws=ws1, column='V9', text='Aplikasi')
#     ws1= help_excel.text_center(ws=ws1, column='W9', text='Before')
#     ws1= help_excel.text_center(ws=ws1, column='X9', text='After')

#     # List Pekerjaan
#     list_pekerjaan = []
#     list_code = []
#     for x in ls_pekerjaan:
#         list_pekerjaan.append(x.pekerjaan_name)
#         list_code.append(x.pekerjaan_code)

#     # insert text miring
#     for row_index, row_data in enumerate([list_pekerjaan], start=8):
#         for col_index, cell_value in enumerate(row_data, start=7):
#             col_name = help_excel.get_col_name(col_index,row_index)
#             ws1 = help_excel.tilt_text(ws=ws1, text=cell_value,column=col_name)
#     # insert data tabel
#     for row_index, row_data in enumerate([list_code], start=9):
#         for col_index, cell_value in enumerate(row_data, start=7):
#             col_name = help_excel.get_col_name(col_index,row_index)
#             ws1 = help_excel.text_center(ws=ws1, text=cell_value,column=col_name)

#     data_excel = []
#     manual_number = 0
#     for x in data:
#         iteration_pekerjaan = 0
#         X1=''
#         X2=''
#         X3=''
#         X4=''
#         X5=''
#         X6=''
#         X7=''
#         X8=''
#         X9=''
#         X10=''
#         for y in x.pm_pekerjaan_log:
#             iteration_pekerjaan += 1
#             if iteration_pekerjaan ==1:
#                 X1=y.pk_asset_status.status_name if y.pk_asset_status else ''
#             elif iteration_pekerjaan ==2:
#                 X2=y.pk_asset_status.status_name if y.pk_asset_status else ''
#             elif iteration_pekerjaan ==3:
#                 X3=y.pk_asset_status.status_name if y.pk_asset_status else ''
#             elif iteration_pekerjaan ==4:
#                 X4=y.pk_asset_status.status_name if y.pk_asset_status else ''
#             elif iteration_pekerjaan ==5:
#                 X5=y.pk_asset_status.status_name if y.pk_asset_status else ''
#             elif iteration_pekerjaan ==6:
#                 X6=y.pk_asset_status.status_name if y.pk_asset_status else ''
#             elif iteration_pekerjaan ==7:
#                 X7=y.pk_asset_status.status_name if y.pk_asset_status else ''
#             elif iteration_pekerjaan ==8:
#                 X8=y.pk_asset_status.status_name if y.pk_asset_status else ''
#             elif iteration_pekerjaan ==9:
#                 X9=y.pk_asset_status.status_name if y.pk_asset_status else ''
#             elif iteration_pekerjaan ==10:
#                 X10=y.pk_asset_status.status_name if y.pk_asset_status else ''
#         manual_number += 1
#         # pengisian data sheet 1
#         xarray = [
#             manual_number,
#             x.id_asset,
#             x.asset.serial_number if x.asset else '',
#             x.asset.name_asset if x.asset else '',
#             x.asset.merek_type if x.asset else '',
#             x.asset.tahun_perangkat if x.asset else '',
#             X1,
#             X2,
#             X3,
#             X4,
#             X5,
#             X6,
#             X7,
#             X8,
#             X9,
#             X10,
#             x.asset.user_eos.name if x.asset else '',
#             x.asset.user_eos.jabatan_user.jabatan_name if x.asset.user_eos.jabatan_user else '',
#             x.asset.user_eos.personal_number if x.asset else '',
#             "ph",
#             "",
#             "√",
#             x.old_foto_label,
#             x.new_foto_label
#         ]
#         data_excel.append(xarray)
#     # Set border halman pertama
#     for row in ws1.iter_rows(min_row=8, max_row=len(data)+9, min_col=1, max_col=24):
#         for cell in row:
#             cell.border = thin_border

#     # tanda tangan
#     for row in ws1.iter_rows(min_row=len(data)+10+12, max_row=len(data)+10+12, min_col=17, max_col=17+1):
#         for cell in row:
#             cell.border = thin_border_up
#     for row in ws1.iter_rows(min_row=len(data)+10+12, max_row=len(data)+10+12, min_col=17+4, max_col=17+4+1):
#         for cell in row:
#             cell.border = thin_border_up
#     ws1 = help_excel.merge_and_left_text(ws=ws1,
#         title=f"{unitkerja_data},",
#         column=help_excel.get_col_name(nomor_baris=len(data)+10+4, nomor_kolom=17),
#         range_col=f'{help_excel.get_col_name(nomor_baris=len(data)+10+4, nomor_kolom=17)}:{help_excel.get_col_name(nomor_baris=len(data)+10+4, nomor_kolom=17+1)}'
#         )            
#     ws1 = help_excel.merge_and_center_text(
#         ws=ws1,title="Kepala Uker / IT KC", 
#         column=help_excel.get_col_name(nomor_baris=len(data)+10+12, nomor_kolom=17), 
#         range_col=f'{help_excel.get_col_name(nomor_baris=len(data)+10+12, nomor_kolom=17)}:{help_excel.get_col_name(nomor_baris=len(data)+10+12, nomor_kolom=17+1)}')
#     ws1 = help_excel.merge_and_left_text(ws=ws1,
#         title="Mengetahui,",
#         column=help_excel.get_col_name(nomor_baris=len(data)+10+4, nomor_kolom=17+4),
#         range_col=f'{help_excel.get_col_name(nomor_baris=len(data)+10+4, nomor_kolom=17+4)}:{help_excel.get_col_name(nomor_baris=len(data)+10+4, nomor_kolom=17+4+1)}'
#         )            
#     ws1 = help_excel.merge_and_center_text(
#         ws=ws1,title="Nama Engineer", 
#         column=help_excel.get_col_name(nomor_baris=len(data)+10+11, nomor_kolom=17+4), 
#         range_col=f'{help_excel.get_col_name(nomor_baris=len(data)+10+11, nomor_kolom=17+4)}:{help_excel.get_col_name(nomor_baris=len(data)+10+11, nomor_kolom=17+4+1)}')
#     ws1 = help_excel.merge_and_center_text(
#         ws=ws1,title="Engineer BriBox", 
#         column=help_excel.get_col_name(nomor_baris=len(data)+10+12, nomor_kolom=17+4), 
#         range_col=f'{help_excel.get_col_name(nomor_baris=len(data)+10+12, nomor_kolom=17+4)}:{help_excel.get_col_name(nomor_baris=len(data)+10+12, nomor_kolom=17+4+1)}')
    
#     # Menentukan baris awal
#     start_row = 10
#     for row_index, row_data in enumerate(data_excel, start=start_row):
#         for col_index, cell_value in enumerate(row_data, start=1):
#             col_name = help_excel.get_col_name(col_index,row_index)
#             ws1 = help_excel.text_center(ws=ws1, text=cell_value,column=col_name)
    
#     # Mulai sheet baru
#     for i in photo_data:
#         print("i[0] : \n",i[0])
#         unique_char = ['/','@', ',']
#         title_sheet=i[0]
#         for j in unique_char:
#             title_sheet = title_sheet.replace(j,' ')
#         title=f"FOTO BUKTI({title_sheet})"
#         print("title : \n",title)
#         ws2 = wb.create_sheet(title=title)
#         ws2 = help_excel.text_right(ws=ws2,text="Nama Perangkat :", column='B3')
#         ws2 = help_excel.text_right(ws=ws2,text="Kode Seri :", column='B4')
#         ws2 = help_excel.text_right(ws=ws2,text="Asset id :", column='B5')
#         ls_label_old_photo = i[1].split(';') if i[1] else []
#         ls_new_label_photo = i[2].split(';') if i[2] else []
#         ls_old_photo = i[3].split(';') if i[3] else []
#         ls_new_photo = i[4].split(';') if i[4] else []
#         # Set border
#         for row in ws2.iter_rows(min_row=2, max_row=11, min_col=2, max_col=len(ls_label_old_photo)+6):
#             for cell in row:
#                 cell.border = thin_border
#         ws2 = help_excel.merge_and_center_text(color=blue_fill, column='B2', range_col=f'B2:{help_excel.get_col_index(nomor_kolom=len(ls_label_old_photo)+6)}2',ws=ws2,title='FOTO BUKTI PM BRIBOX')
#         ws2 = help_excel.merge_and_left_text(ws=ws2,column='C3',range_col=f'C3:{help_excel.get_col_index(nomor_kolom=len(ls_label_old_photo)+6)}3',title=i.nama_asset)
#         ws2 = help_excel.merge_and_left_text(ws=ws2,column='C4',range_col=f'C4:{help_excel.get_col_index(nomor_kolom=len(ls_label_old_photo)+6)}4',title="")
#         ws2 = help_excel.merge_and_left_text(ws=ws2,column='C5',range_col=f'C5:{help_excel.get_col_index(nomor_kolom=len(ls_label_old_photo)+6)}5',title="")
#         ws2 = help_excel.merge_and_center_text(ws=ws2, range_col=f'B6:{help_excel.get_col_index(nomor_kolom=len(ls_label_old_photo)+6)}6', column='B6', title='SEBELUM', color=blue_fill)
#         ws2 = help_excel.merge_and_center_text(ws=ws2, range_col=f'B9:{help_excel.get_col_index(nomor_kolom=len(ls_label_old_photo)+6)}9', column='B9', title='SESUDAH', color=blue_fill)
#         # Menentukan baris awal
#         start_row = 7
#         for row_index, row_data in enumerate([ls_label_old_photo], start=start_row):
#             for col_index, cell_value in enumerate(row_data, start=2):
#                 col_name = help_excel.get_col_name(col_index,row_index)
#                 ws2 = help_excel.text_center(ws=ws2, text=cell_value,column=col_name)
#         start_row = 8
#         for row_index, row_data in enumerate([ls_old_photo], start=start_row):
#             for col_index, cell_value in enumerate(row_data, start=2):
#                 col_name = help_excel.get_col_name(col_index,row_index)
#                 # ws2 = insert_gambar(ws=ws2, column=col_name,column_id=get_col_index(col_index),row=row_index)
#                 ws2 = help_excel.insert_gambar_custom(ws=ws2, column=col_name,column_id=help_excel.get_col_index(col_index),row=row_index, path_file=cell_value)
#         start_row = 10
#         for row_index, row_data in enumerate([ls_new_label_photo], start=start_row):
#             for col_index, cell_value in enumerate(row_data, start=2):
#                 col_name = help_excel.get_col_name(col_index,row_index)
#                 ws2 = help_excel.text_center(ws=ws2, text=cell_value,column=col_name)
#         start_row = 11
#         for row_index, row_data in enumerate([ls_new_photo], start=start_row):
#             for col_index, cell_value in enumerate(row_data, start=2):
#                 col_name = help_excel.get_col_name(col_index,row_index)
#                 ws2 = help_excel.insert_gambar_custom(ws=ws2, column=col_name,column_id=help_excel.get_col_index(col_index),row=row_index, path_file=cell_value)
    
#     # Delete default ws
#     del wb['Sheet']
#     # Save workbook in memory
#     with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
#         wb.save(tmp.name)
#         tmp.seek(0)
#         upload_file_file = UploadFile(filename="data.xlsx", file=io.BytesIO(tmp.read()))
#         return upload_file_file