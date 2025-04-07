'''
this function will generate a report of all the clients and their billing information
this will use kiss methods to keep the code clean and simple
this will use RafiExcel class to generate the excel file
'''
import io
from tempfile import NamedTemporaryFile
from core.rafiexcel import RafiExcel
from models.EmployeeAllowances import EmployeeAllowances
from models.EmployeeTax import EmployeeTax
from models.BpjsEmployee import BpjsEmployee
from models.ClientPayment import ClientPayment
from models.Attendance import Attendance
from models.Payroll import Payroll
from models.Client import Client
from models import SessionLocal
from datetime import date, datetime, timedelta
from settings import TZ
import calendar
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
from starlette.datastructures import UploadFile
from core.file import upload_file


rafi_excel = RafiExcel()

def format_to_idr(value):
    return f"Rp {value:,.2f}".replace(',', '.').replace('.', ',', 1)

async def generate_client_billing_report(id_client: int):
    '''
    this function will generate a report of all the clients and their billing information
    this will use kiss methods to keep the code clean and simple
    this will use RafiExcel class to generate the excel file
    this will use SessionLocal to get the data from the database
    '''
    db = SessionLocal()
    try:
        # get data client
        client = db.query(Client).filter(Client.id == id_client).first()

        if not client:
            raise ValueError("Client not found")
        # get data periode data
        client_payment_date = client.due_date_payment.strftime("%d")
        today = datetime.now().date()
        start_periode = today.replace(day=int(client_payment_date) - 1)
        end_periode = (start_periode + timedelta(days=30)).replace(day=int(client_payment_date) - 1)

        # get data start and end date this month
        start_date = datetime.now().date().replace(day=1)
        _, last_day_of_month = calendar.monthrange(start_date.year, start_date.month)
        end_date = start_date.replace(day=last_day_of_month) + timedelta(days=1)

        # get data client payment
        client_payment = db.query(ClientPayment).filter(
            ClientPayment.client_id == id_client,
            ClientPayment.date <= end_date,
            ClientPayment.date >= start_date,
            ClientPayment.isact == True,
        ).first()

        # get data employee payroll
        payroll = db.query(Payroll).filter(
            Payroll.client_id == id_client,
            Payroll.payment_date <= end_date,
            Payroll.payment_date >= start_date,
            Payroll.isact == True,
        ).all()

        # get data employee tax
        employee_tax = db.query(EmployeeTax).filter(
            EmployeeTax.client_id == id_client,
            EmployeeTax.isact == True,
        ).all()
        
        # get data employee allowances
        employee_allowances = db.query(EmployeeAllowances).filter(
            EmployeeAllowances.client_id == id_client,
            EmployeeAllowances.isact == True,
        ).all()

        # get data employee bpjs
        employee_bpjs = db.query(BpjsEmployee).filter(
            BpjsEmployee.client_id == id_client,
            BpjsEmployee.isact == True,
        ).all()

        # Data preparation for the report
        tittle_lk_karyawan = "Laporan Pengeluran Biaya"
        client_name = client.name if client else "Unknown Client"
        periode = f"{start_periode.strftime('%d %B %Y')} - {end_periode.strftime('%d %B %Y')}"
        biaya_operasional = sum([p.monthly_paid for p in payroll if p.monthly_paid]) if payroll else 0
        biaya_agency = (client.fee_agency / 100) * biaya_operasional if client.fee_agency else 0
        biaya_allowances = sum([e.amount for e in employee_allowances if e.amount]) if employee_allowances else 0 
        biaya_bpjs = sum([b.amount for b in employee_bpjs if b.amount]) if BpjsEmployee else 0
        biaya_tax = sum([t.amount for t in employee_tax if t.amount]) if employee_tax else 0
        grand_total = biaya_operasional + biaya_agency + biaya_allowances + biaya_bpjs + biaya_tax

        # generate sheet lk_karyawan
        wb = Workbook()
        ws1 = wb.create_sheet(title='lk_kayawan', index=0)
        ws1 = rafi_excel.merge_and_center_text(column='A1', range_col='A1:C1', ws=ws1, title=client_name)
        ws1 = rafi_excel.merge_and_center_text(column='B3', range_col='B3:F3', ws=ws1, title=tittle_lk_karyawan)
        ws1 = rafi_excel.merge_and_center_text(column='B4', range_col='B4:F4', ws=ws1, title=client_name)
        ws1 = rafi_excel.merge_and_center_text(column='B5', range_col='B5:F5', ws=ws1, title=periode)

        # Fill header
        data_header = [
            "No",
            "Keterangan",
            "Rp.",
            "Nominal",
            "Jumlah",
        ]
        index_header = 0
        for item in data_header:
            index_header += 1
            column_name = rafi_excel.get_col_name(nomor_kolom=index_header+1, nomor_baris=8)
            ws1 = rafi_excel.text_center(ws=ws1, column=column_name, text=item)
        
        data_table = [
            {'keterangan': 'Biaya Operasional', 'nominal': format_to_idr(biaya_operasional), 'rp':'Rp.', 'jumlah': None},
            {'keterangan': 'Biaya Allowances', 'nominal': format_to_idr(biaya_allowances), 'rp':'Rp.', 'jumlah': None},
            {'keterangan': 'Biaya BPJS', 'nominal': format_to_idr(biaya_bpjs), 'rp':'Rp.', 'jumlah': None},
            {'keterangan': 'Jumlah biaya operasional', 'nominal': None, 'rp':'Rp.', 'jumlah': format_to_idr(biaya_operasional+biaya_allowances+biaya_bpjs)},
            {'keterangan': 'Agency Fee', 'nominal': format_to_idr(biaya_agency), 'rp':'Rp.', 'jumlah': None},
            {'keterangan': 'Jumlah Biaya', 'nominal': None, 'rp':'Rp.', 'jumlah': format_to_idr(biaya_agency)},
            {'keterangan': 'Biaya Pajak', 'nominal': format_to_idr(biaya_tax), 'rp':'Rp.', 'jumlah': None},
            {'keterangan': 'Grand Total', 'nominal': None, 'rp':'Rp.', 'jumlah': format_to_idr(grand_total)},
        ]
        index_table = 0
        start_index_table = 8
        for item in data_table:
            index_table += 1
            column_name = rafi_excel.get_col_name(nomor_kolom=2, nomor_baris=index_table + start_index_table)
            ws1 = rafi_excel.text_center(ws=ws1, column=column_name, text=index_table)

            column_name = rafi_excel.get_col_name(nomor_kolom=3, nomor_baris=index_table + start_index_table)
            if not item['nominal']:
                ws1 = rafi_excel.text_center_bold(ws=ws1, column=column_name, text=item['keterangan'])
            ws1 = rafi_excel.text_center(ws=ws1, column=column_name, text=item['keterangan'])

            column_name = rafi_excel.get_col_name(nomor_kolom=4, nomor_baris=index_table + start_index_table)
            ws1 = rafi_excel.text_center(ws=ws1, column=column_name, text=item['rp'])

            column_name = rafi_excel.get_col_name(nomor_kolom=5, nomor_baris=index_table + start_index_table)
            if not item['nominal']:
                ws1 = rafi_excel.text_center_bold(ws=ws1, column=column_name, text=item['nominal'])
            ws1 = rafi_excel.text_center(ws=ws1, column=column_name, text=item['nominal'])

            column_name = rafi_excel.get_col_name(nomor_kolom=6, nomor_baris=index_table + start_index_table)
            if not item['nominal']:
                ws1 = rafi_excel.text_center_bold(ws=ws1, column=column_name, text=item['jumlah'])
            ws1 = rafi_excel.text_center(ws=ws1, column=column_name, text=item['jumlah'])

        # set border
        all_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws1.iter_rows(min_row=8, max_row=index_table + 8, min_col=2, max_col=6):
            for cell in row:
                cell.border = all_border

        # set signature column
        start_index_signature = index_table + start_index_table + 2
        column_name_start = rafi_excel.get_col_name(nomor_kolom=1, nomor_baris=start_index_signature)
        column_name_end = rafi_excel.get_col_name(nomor_kolom=6, nomor_baris=start_index_signature)
        ws1 = rafi_excel.merge_and_center_text(column=f'A{start_index_signature}', range_col=f'{column_name_start}:{column_name_end}', ws=ws1, title=client_name)
        ws1 = rafi_excel.text_center(ws=ws1, column=f'A{start_index_signature+1}', text="Dibuat oleh")
        ws1 = rafi_excel.merge_and_center_text(column=f'B{start_index_signature+1}', range_col=f'B{start_index_signature+1}:C{start_index_signature+1}', ws=ws1, title="Diperiksa Oleh")
        ws1 = rafi_excel.merge_and_center_text(column=f'D{start_index_signature+1}', range_col=f'D{start_index_signature+1}:E{start_index_signature+1}', ws=ws1, title="Diketahui Oleh")
        ws1 = rafi_excel.text_center(ws=ws1, column=f'F{start_index_signature+1}', text="Disetujui Oleh")
        
        #set hight column for sigature
        ws1.row_dimensions[start_index_signature+2].height = 50

        # set name in signature column
        ws1 = rafi_excel.text_center_bold(ws=ws1, column=f'A{start_index_signature+3}', text="Payroll")
        ws1 = rafi_excel.text_center_bold(ws=ws1, column=f'B{start_index_signature+3}', text="Finance")
        ws1 = rafi_excel.text_center_bold(ws=ws1, column=f'C{start_index_signature+3}', text="Finance")
        ws1 = rafi_excel.text_center_bold(ws=ws1, column=f'D{start_index_signature+3}', text="HRD")
        ws1 = rafi_excel.text_center_bold(ws=ws1, column=f'F{start_index_signature+3}', text="Direktur")


        # set column width
        ws1.column_dimensions['C'].width = 40
        ws1.column_dimensions['E'].width = 20
        ws1.column_dimensions['F'].width = 20

        # Delete default ws
        del wb['Sheet']
        # Save workbook in memory
        formated_date_payment = client.due_date_payment.strftime("%d-%m-%Y")
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            upload_file_file = UploadFile(filename="data.xlsx", file=io.BytesIO(tmp.read()))
            # return upload_file_file
            new_path = await upload_file(
                upload_file=upload_file_file, path = f"/Client_Payment/{formated_date_payment}/Data_payment.xlsx"
            )
    except Exception as e:
        print(f"Error: {e}")
        raise ValueError("Error generating report", e)
    finally:
        db.close()

async def generate_employee_attendance(id_client: int):
    '''
    This function generates a report of all employee attendance.
    It uses the RafiExcel class to generate the Excel file and SessionLocal to fetch data from the database.
    '''
    db = SessionLocal()
    try:
        # Get client data
        client = db.query(Client).filter(Client.id == id_client).first()
        if not client:
            raise ValueError("Client not found")

        # Get period data
        client_payment_date = client.due_date_payment.strftime("%d")
        today = datetime.now().date()
        start_periode = today.replace(day=int(client_payment_date) - 1)
        end_periode = (start_periode + timedelta(days=30)).replace(day=int(client_payment_date) - 1)

        # Get attendance data based on the period
        attendance = db.query(Attendance).filter(
            Attendance.client_id == id_client,
            Attendance.date <= end_periode,
            Attendance.date >= start_periode,
            Attendance.isact == True,
        ).order_by(Attendance.date).all()

        # Group attendance by date
        attendance_by_date = {}
        for record in attendance:
            date_key = record.date.strftime('%Y-%m-%d')
            if date_key not in attendance_by_date:
                attendance_by_date[date_key] = []
            attendance_by_date[date_key].append(record)

        # Create day and date headers using the calendar library
        day_date_header = []
        current_date = start_periode
        while current_date <= end_periode:
            day_name = calendar.day_name[current_date.weekday()][:3]  # Shortened day name (Mon, Tue, etc.)
            day_date = f"{day_name}\n{current_date.day}"
            day_date_header.append(day_date)
            current_date += timedelta(days=1)

        # Generate Excel report
        wb = Workbook()
        ws1 = wb.create_sheet(title='Attendance', index=0)

        # Create headers
        ws1 = rafi_excel.merge_and_center_text(column='A1', range_col='A1:C1', ws=ws1, title=client.name)
        ws1 = rafi_excel.merge_and_center_text(column='B3', range_col='B3:F3', ws=ws1, title="Employee Attendance Report")
        ws1 = rafi_excel.merge_and_center_text(column='B4', range_col='B4:F4', ws=ws1, title=f"Period: {start_periode.strftime('%d %B %Y')} - {end_periode.strftime('%d %B %Y')}")

        # Set column headers
        ws1.cell(row=6, column=1).value = "No"
        ws1.cell(row=6, column=2).value = "Employee Name"

        # Add day/date headers
        for idx, day_date in enumerate(day_date_header):
            col = idx + 3
            cell = ws1.cell(row=6, column=col)
            cell.value = day_date
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws1.column_dimensions[rafi_excel.get_col_index(col)].width = 8

        # Add total column
        total_col = len(day_date_header) + 3
        ws1.cell(row=6, column=total_col).value = "Total"

        # Get unique employees from attendance
        employees = {}
        for date_records in attendance_by_date.values():
            for record in date_records:
                if record.emp_id not in employees:
                    employees[record.emp_id] = record.users.name

        # Fill attendance data
        row = 7
        for idx, (employee_id, employee_name) in enumerate(sorted(employees.items(), key=lambda x: x[1])):
            ws1.cell(row=row, column=1).value = idx + 1
            ws1.cell(row=row, column=2).value = employee_name

            # Initialize total attendance for this employee
            total_present = 0

            # Fill attendance status for each date
            current_date = start_periode
            col_idx = 3
            while current_date <= end_periode:
                date_key = current_date.strftime('%Y-%m-%d')
                attendance_status = ''

                # Check if employee has attendance on this date
                if date_key in attendance_by_date:
                    for record in attendance_by_date[date_key]:
                        if record.emp_id == employee_id:
                            attendance_status = 'P'  # Present
                            total_present += 1
                            break

                # If no attendance found and it's a weekday, mark as absent
                if not attendance_status and current_date.weekday() < 5:  # 0-4 is Monday to Friday
                    attendance_status = 'A'  # Absent

                # Weekend gets special marking
                if not attendance_status and current_date.weekday() >= 5:  # 5-6 is Saturday and Sunday
                    attendance_status = 'W'  # Weekend

                ws1.cell(row=row, column=col_idx).value = attendance_status
                ws1.cell(row=row, column=col_idx).alignment = Alignment(horizontal='center')

                current_date += timedelta(days=1)
                col_idx += 1

            # Add total for this employee
            ws1.cell(row=row, column=total_col).value = total_present
            ws1.cell(row=row, column=total_col).alignment = Alignment(horizontal='center')
            row += 1

        # Apply borders
        all_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row_cells in ws1.iter_rows(min_row=6, max_row=row - 1, min_col=1, max_col=total_col):
            for cell in row_cells:
                cell.border = all_border

        # Set column width
        ws1.column_dimensions['B'].width = 30

        # Delete default sheet
        del wb['Sheet']

        # Save workbook
        formated_date_payment = client.due_date_payment.strftime("%d-%m-%Y")
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            upload_file_file = UploadFile(filename="attendance.xlsx", file=io.BytesIO(tmp.read()))
            new_path = await upload_file(
                upload_file=upload_file_file, path=f"/Client_Payment/{formated_date_payment}/Employee_Attendance.xlsx"
            )

        return new_path

    except Exception as e:
        print(f"Error: {e}")
        raise ValueError("Error generating report", e)
    finally:
        db.close()