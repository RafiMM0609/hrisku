"""
This file need sets to user KIS (keep it simple) method
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
import calendar
from core.rafiexcel import RafiExcel, blue_fill, yellow_fill
from models import SessionLocal
from sqlalchemy import select, func, or_
from sqlalchemy.orm import Session, joinedload
from datetime import datetime, timedelta, date
from models.Client import Client
from models.Payroll import Payroll
from models.EmployeeAllowances import EmployeeAllowances
from models.EmployeeTax import EmployeeTax
from models.BpjsEmployee import BpjsEmployee
from models.User import User
from tempfile import NamedTemporaryFile
from starlette.datastructures import UploadFile
from core.file import upload_file, upload_file_to_tmp
import io

def get_days_in_month(year, month):
    return calendar.monthrange(year, month)[1]

# # Contoh penggunaan
# year = 2025
# month = 4

# days_in_month = get_days_in_month(year, month)
# print(f"Bulan {month}/{year} memiliki {days_in_month} hari.")


async def generate_file_excel(emp_id, client_id):
    '''
    this function will generate file excel for employee payroll
    this function will user LocalSession for database connection
    this function will return file path for the generated file
    excel will have dynamic heading based on the data on UserAllowence, UserTax, UserBPJS
    this excel will generated every month for the employee
    '''
    db = SessionLocal()
    try:
        start_of_this_month = datetime(datetime.now().year, datetime.now().month, 1)
        end_of_this_month = datetime(datetime.now().year, datetime.now().month, calendar.monthrange(datetime.now().year, datetime.now().month)[1])
        
        emp_data = db.query(User)\
            .filter(
                User.client_id == client_id,
                User.id == emp_id,
                User.isact == True
            )\
            .first()

        emp_payroll = db.query(Payroll)\
            .filter(
                Payroll.client_id == client_id,
                Payroll.emp_id == emp_id,
                Payroll.isact == True,
                Payroll.payment_date.between(start_of_this_month.date(), end_of_this_month.date()),  # Updated filter
            )\
            .first()
        if not emp_payroll:
            raise ValueError("Employee contract not found for the given client.")
        emp_allowances = db.query(EmployeeAllowances)\
            .filter(
                EmployeeAllowances.client_id == client_id,
                EmployeeAllowances.emp_id == emp_id,
                EmployeeAllowances.isact == True,
                EmployeeAllowances.created_at.between(start_of_this_month, end_of_this_month),  # Updated filter
            )\
            .all()
        emp_tax = db.query(EmployeeTax)\
            .filter(
                EmployeeTax.client_id == client_id,
                EmployeeTax.emp_id == emp_id,
                EmployeeTax.isact == True,
                EmployeeTax.created_at.between(start_of_this_month, end_of_this_month),  # Updated filter
            )\
            .all()
        emp_bpjs = db.query(BpjsEmployee)\
            .filter(
                BpjsEmployee.client_id == client_id,
                BpjsEmployee.emp_id == emp_id,
                BpjsEmployee.isact == True,
                BpjsEmployee.created_at.between(start_of_this_month, end_of_this_month),  # Updated filter
            )\
            .all()
        
        def format_to_idr(value):
            return f"Rp {value:,.2f}".replace(',', '.').replace('.', ',', 1)
        #  Create excel data
        def generate_excel_data(emp_payroll:Payroll, emp_allowances:EmployeeAllowances, emp_tax:EmployeeTax, emp_bpjs:BpjsEmployee):
            fromated_date_payment = emp_payroll.payment_date.strftime("%d %B %Y")

            help_excel = RafiExcel()
            wb = Workbook()
            ws1 = wb.create_sheet(title='Employee Payroll')
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col='A1:C1', column='A1',color=blue_fill, title="Name of Employee")
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col='A2:C2', column='A2',color=blue_fill, title=f"{emp_data.name}")
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col='A4:J4', column='A4', title="Payslip")
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col='A5:J5', column='A5', title=f"{fromated_date_payment}")

            # Setup Income
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col='A9:C9', column='A9', title="Total Income")
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col='E9:F9', column='E9', title="IDR")

            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col='A10:C10', column='A10', title="Salary")
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col='E10:F10', column='E10', title="IDR")
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col='H10:J10', column='H10', title=f"{format_to_idr(emp_payroll.monthly_paid)}")
            
            index_allowance = 1
            total_allowance = 0.00
            for item in emp_allowances:
                ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'A{10+index_allowance}:C{10+index_allowance}', column=f'A{10+index_allowance}', title=item.name)
                ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'E{10+index_allowance}:F{10+index_allowance}', column=f'E{10+index_allowance}', title="IDR")
                ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'H{10+index_allowance}:J{10+index_allowance}', column=f'H{10+index_allowance}', title=f"{format_to_idr(item.amount)}")
                index_allowance += 1
                total_allowance += item.amount
            
            total_income = emp_payroll.monthly_paid + total_allowance
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col='H9:J9', column='H9', title=f'{format_to_idr(total_income)}')

            index_allowance += 1
            # Setup Deduction
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'A{10+index_allowance}:C{10+index_allowance}', column=f'A{10+index_allowance}', title="Total Deduction")
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'E{10+index_allowance}:F{10+index_allowance}', column=f'E{10+index_allowance}', title="IDR")

            total_deduction = 0.00
            index_deduction = 1
            for item in emp_bpjs:
                ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'A{10+index_allowance+index_deduction}:C{10+index_allowance+index_deduction}', column=f'A{10+index_allowance+index_deduction}', title=item.name)
                ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'E{10+index_allowance+index_deduction}:F{10+index_allowance+index_deduction}', column=f'E{10+index_allowance+index_deduction}', title="IDR")
                ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'H{10+index_allowance+index_deduction}:J{10+index_allowance+index_deduction}', column=f'H{10+index_allowance+index_deduction}', title=f"{format_to_idr(item.amount)}")
                index_deduction += 1
                total_deduction += item.amount
            index_tax = 1
            for item in emp_tax:
                ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'A{10+index_allowance+index_deduction+index_tax}:C{10+index_allowance+index_deduction+index_tax}', column=f'A{10+index_allowance+index_deduction+index_tax}', title=item.name)
                ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'E{10+index_allowance+index_deduction+index_tax}:F{10+index_allowance+index_deduction+index_tax}', column=f'E{10+index_allowance+index_deduction+index_tax}', title="IDR")
                ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'H{10+index_allowance+index_deduction+index_tax}:J{10+index_allowance+index_deduction+index_tax}', column=f'H{10+index_allowance+index_deduction+index_tax}', title=f"{format_to_idr(item.amount)}")
                index_tax += 1
                total_deduction += item.amount

            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'H{10+index_allowance}:J{10+index_allowance}', column=f'H{10+index_allowance}', title=f"{format_to_idr(total_deduction)}")

            # Setup Net Salary
            latest_index = 10 + index_allowance + index_deduction + index_tax
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'A{latest_index}:C{latest_index}', column=f'A{latest_index}', title="Net Salary")
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'E{latest_index}:F{latest_index}', column=f'E{latest_index}', title="IDR")
            ws1 = help_excel.merge_and_center_text(ws=ws1,range_col=f'H{latest_index}:J{latest_index}', column=f'H{latest_index}', title=f"{format_to_idr(emp_payroll.net_salary)}")


            # Delete default ws
            del wb['Sheet']
            # Save workbook in memory
            with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                upload_file_file = UploadFile(filename="data.xlsx", file=io.BytesIO(tmp.read()))
                return upload_file_file
        excel_file = generate_excel_data(emp_payroll, emp_allowances, emp_tax, emp_bpjs)
        formated_date_payment = emp_payroll.payment_date.strftime("%d-%m-%Y")
        # path = await upload_file_to_tmp(
        #     upload_file=excel_file, filename = f"Data-Payroll-{emp_data.name}-{formated_date_payment}.xlsx"
        # )
        new_path = await upload_file(
            upload_file=excel_file, path = f"/Payroll/{formated_date_payment}/Data-Payroll-{emp_data.name}-{formated_date_payment}.xlsx"
        )
        emp_payroll.file = new_path
        db.add(emp_payroll)
        db.commit()
        return new_path
    except Exception as e:
        print("Error generate excel file: \n", e)
        raise ValueError("Failed to generate excel file: \n", e)
    finally:
        db.close()


async def add_monthly_salary_emp(emp_id, client_id):
    db = SessionLocal()
    try:
        # Fetch the end date of the contract from ContractClient
        # contract = (
        #     db.query(ContractClient)
        #     .filter(ContractClient.client_id == client_id)
        #     .order_by(ContractClient.end.desc())
        #     .first()
        # )
        # if not contract or not contract.end:
        #     raise ValueError("Contract end date not found for the given client.")
        
        # Data preparation
        start_of_year = date(date.today().year, 1, 1)
        end_of_year = date(date.today().year, 12, 31)

        # Calculate the start and end of the current month
        start_of_month = datetime(datetime.now().year, datetime.now().month, 1)
        end_of_month = datetime(datetime.now().year, datetime.now().month, calendar.monthrange(datetime.now().year, datetime.now().month)[1])
        # Payment_date
        payment_date = end_of_month - timedelta(days=1)  # Generate dummy date 30 days from today
        
        # Query ClientPayment dengan eager loading untuk menghindari query tambahan
        client = db.query(Client)\
            .options(
            joinedload(Client.client_tax),
            joinedload(Client.allowances),
            joinedload(Client.bpjs),
            joinedload(Client.user_client)
            )\
            .filter(Client.isact == True, Client.id == client_id)\
            .first()
        # client = db.execute(query).scalar_one_or_none()
        taxes = client.client_tax
        allowances = client.allowances
        bpjs_data = client.bpjs
        # Initialize total nominal with basic salary
        total_nominal = 0.00
        basic_salary = client.basic_salary or 0.00
        total_nominal += basic_salary

        # Perhitungan pergaji karyawan

        # Add BPJS
        ls_bpjs = []
        ls_allowances = []
        ls_tax = []
        total_bpjs = 0.00
        total_allowances = 0.00
        total_tax = 0.00

        # Add BPJS
        for item in bpjs_data:
            bpjs_amount = ((item.amount or 0) / 100) * basic_salary
            ls_bpjs.append(
                {
                    "name": item.name,
                    "amount": bpjs_amount,
                }
            )
            total_bpjs += bpjs_amount

        # Add Allowances
        for item in allowances:
            allowance_amount = item.amount or 0
            ls_allowances.append(
                {
                    "name": item.name,
                    "amount": allowance_amount,
                }
            )
            total_allowances += allowance_amount
            total_nominal += allowance_amount

        # Hitung total gaji dalam seperiode
        count_payroll = db.execute(
            select(func.count(Payroll.id))
            .filter(
                Payroll.client_id == client_id,
                Payroll.emp_id == emp_id,
                Payroll.isact == True,
                Payroll.payment_date.between(start_of_year, end_of_year),  # Updated filter
            )
        ).scalar()
        if count_payroll >= 11:
            # Add tax
            for item in taxes:
                tax_amount = ((item.percent or 0) / 100) * (basic_salary * 12)
                ls_tax.append(
                    {
                        "name": item.name,
                        "amount": tax_amount,
                    }
                )

        # Check payroll bulan ini sudah ada belum
        exist_payroll = db.query(Payroll)\
            .filter(
                Payroll.emp_id==emp_id, 
                Payroll.isact==True,
                Payroll.payment_date.between(start_of_month, end_of_month)  # Updated filter
                )\
            .first()

        # Count net sallary
        net_salary = total_nominal - total_bpjs

        # Add or update allowances record
        for item in ls_allowances:
            existing_allowance = db.query(EmployeeAllowances)\
                .filter(
                    EmployeeAllowances.client_id == client_id,
                    EmployeeAllowances.emp_id == emp_id,
                    EmployeeAllowances.name == item["name"],
                    func.date(EmployeeAllowances.created_at) >= start_of_month.date(),
                    func.date(EmployeeAllowances.created_at) <= end_of_month.date()
                ).first()
            if existing_allowance:
                existing_allowance.amount = item["amount"]
                existing_allowance.created_at = datetime.now()
                db.add(existing_allowance)
            else:
                new_allowance = EmployeeAllowances(
                    client_id=client_id,
                    emp_id=emp_id,
                    name=item["name"],
                    amount=item["amount"],
                    created_at=datetime.now(),
                )
                db.add(new_allowance)

        # Add or update BPJS record
        for item in ls_bpjs:
            existing_bpjs = db.query(BpjsEmployee)\
                .filter(
                    BpjsEmployee.client_id == client_id,
                    BpjsEmployee.emp_id == emp_id,
                    BpjsEmployee.name == item["name"],
                    func.date(BpjsEmployee.created_at) >= start_of_month.date(),
                    func.date(BpjsEmployee.created_at) <= end_of_month.date()
                ).first()
            if existing_bpjs:
                existing_bpjs.amount = item["amount"]
                existing_bpjs.created_at = datetime.now()
                db.add(existing_bpjs)
            else:
                new_bpjs = BpjsEmployee(
                    client_id=client_id,
                    emp_id=emp_id,
                    name=item["name"],
                    amount=item["amount"],
                    created_at=datetime.now(),
                )
                db.add(new_bpjs)

        # Add or update tax record
        for item in ls_tax:
            existing_tax = db.query(EmployeeTax)\
                .filter(
                    EmployeeTax.client_id == client_id,
                    EmployeeTax.emp_id == emp_id,
                    EmployeeTax.name == item["name"],
                    func.date(EmployeeTax.created_at) >= start_of_month.date(),
                    func.date(EmployeeTax.created_at) <= end_of_month.date()
                ).first()
            if existing_tax:
                existing_tax.amount = item["amount"]
                existing_tax.created_at = datetime.now()
                db.add(existing_tax)
            else:
                new_tax = EmployeeTax(
                    client_id=client_id,
                    emp_id=emp_id,
                    name=item["name"],
                    amount=item["amount"],
                    created_at=datetime.now(),
                )
                db.add(new_tax)

        if exist_payroll:
            exist_payroll.monthly_paid = basic_salary
            exist_payroll.payment_date = payment_date
            exist_payroll.total_allowances = total_allowances
            exist_payroll.total_bpjs = total_bpjs
            exist_payroll.total_tax = total_tax
            exist_payroll.net_salary = net_salary
            db.add(exist_payroll)
        else:
            # Add new payroll record
            new_payment = Payroll(
                client_id=client_id,
                emp_id=emp_id,
                monthly_paid=basic_salary,
                payment_date=payment_date,
                total_allowances=total_allowances,
                total_bpjs=total_bpjs,
                total_tax=total_tax,
                net_salary=net_salary,
            )
            db.add(new_payment)
        db.commit()
        return "oke"
    except Exception as e:
        print("Error add user payroll: \n", e)
        raise ValueError("Failed to add user payroll: \n", e)
    finally:
        db.close()

